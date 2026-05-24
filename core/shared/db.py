# core/db.py
"""뉴스 DB 누적 저장 모듈 (xlsx)"""

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

DB_PATH = Path("storage/news_db.xlsx")
SHEET_NAME = "뉴스DB"

_HEADERS = ["날짜", "카테고리", "출처라벨", "언어", "제목", "링크", "발행일시", "요약"]
_COL_KEYS = ["date", "category", "label", "lang", "title", "link", "published", "summary"]
_COL_WIDTHS = [12, 20, 16, 6, 60, 80, 22, 45]
_LINK_COL_IDX = 6  # 1-based, link 컬럼 위치


def _create_sheet(wb: openpyxl.Workbook) -> openpyxl.worksheet.worksheet.Worksheet:
    ws = wb.create_sheet(title=SHEET_NAME)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    for col_idx, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    return ws


def _load_workbook() -> tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet]:
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    if DB_PATH.exists():
        wb = openpyxl.load_workbook(DB_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else _create_sheet(wb)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = _create_sheet(wb)
    return wb, ws


def _existing_links(ws) -> set:
    return {row[_LINK_COL_IDX - 1] for row in ws.iter_rows(min_row=2, values_only=True) if row[_LINK_COL_IDX - 1]}


def append_news(news_items: list, date_str: str) -> int:
    """
    news_items: collect_news()["all"] 리스트
    date_str:   "YYYY-MM-DD"
    반환:       실제로 추가된 행 수
    """
    import os
    if os.getenv("GITHUB_ACTIONS") == "true":
        logger.info("[DB 저장] GitHub Actions 환경 — 로컬 xlsx DB 쓰기 생략 (reports-data.json 캐시 필터로 대체)")
        return 0

    if not news_items:
        return 0

    wb, ws = _load_workbook()
    seen = _existing_links(ws)
    added = 0

    for item in news_items:
        link = item.get("link", "")
        if not link or link in seen:
            continue
        row = [date_str] + [item.get(k, "") for k in _COL_KEYS if k != "date"]
        ws.append(row)
        seen.add(link)
        added += 1

    if added:
        wb.save(DB_PATH)
        logger.info(f"[DB 저장] {added}건 추가 → {DB_PATH}")
    else:
        logger.info("[DB 저장] 신규 항목 없음 (모두 중복)")

    return added
